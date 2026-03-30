import os
import json
import re
import shutil
from datetime import datetime, timedelta
from collections import defaultdict

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import kif_download
    KIF_DOWNLOAD_AVAILABLE = True
except ImportError:
    KIF_DOWNLOAD_AVAILABLE = False

WARS_ID = os.getenv("WARS_ID", "")


def _get_opponent_style(kif_text: str, wars_id: str = WARS_ID) -> str:
    """raw.kif のヘッダーから相手の戦型（最も左の戦法名）を取得する。
    ユーザーが先手のとき → 「後手の戦法：」を参照
    ユーザーが後手のとき → 「先手の戦法：」を参照
    """
    is_sente = False
    for line in kif_text.splitlines():
        if f"先手：{wars_id}" in line or f"先手: {wars_id}" in line:
            is_sente = True
            break
        if f"後手：{wars_id}" in line or f"後手: {wars_id}" in line:
            is_sente = False
            break

    target_prefix = "後手の戦法：" if is_sente else "先手の戦法："

    for line in kif_text.splitlines():
        if line.startswith(target_prefix):
            #style = line[len(target_prefix):].split(",")[0].strip()
            style = line[len(target_prefix):].split(",")[-1].strip()
            return style if style else "その他"
    return "その他"


def _is_user_lost(kif_text: str, wars_id: str = WARS_ID) -> bool:
    """raw.kif の内容からユーザーが負けたか判定する"""
    is_sente = False
    is_gote = False
    for line in kif_text.splitlines():
        if f"先手：{wars_id}" in line or f"先手: {wars_id}" in line:
            is_sente = True
        if f"後手：{wars_id}" in line or f"後手: {wars_id}" in line:
            is_gote = True

    sente_won = (
        "先手の勝ち" in kif_text
        or "先手勝ち" in kif_text
        or "勝者：▲" in kif_text
    )
    gote_won = (
        "後手の勝ち" in kif_text
        or "後手勝ち" in kif_text
        or "勝者：△" in kif_text
    )

    if is_sente:
        return gote_won   # 先手がユーザー → 後手が勝ったら負け
    elif is_gote:
        return sente_won  # 後手がユーザー → 先手が勝ったら負け
    return False


def _collect_battle_kif_paths(ROOT: str) -> dict:
    """temp, backup, 戦型別 配下を再帰走査し、
    battle_id → raw.kif パスの辞書を返す（重複は最初に見つかったものを使用）"""
    search_dirs = [
        os.path.join(ROOT, "temp"),
        os.path.join(ROOT, "backup"),
        os.path.join(ROOT, "戦型別"),
    ]
    battles = {}
    for base_dir in search_dirs:
        if not os.path.exists(base_dir):
            continue
        for root_dir, dirs, files in os.walk(base_dir):
            for d in dirs:
                if re.match(r'.+-\d{8}_\d{6}$', d):
                    kif_path = os.path.join(root_dir, d, "raw.kif")
                    if os.path.exists(kif_path) and d not in battles:
                        battles[d] = kif_path
    return battles


def _build_rate_row(ROOT: str, log_path: str):
    """rate_trg ディレクトリを作成し、100件ダウンロード→1ヶ月以前を削除→集計して
    1行レコード [集計日付, 対局数, 勝利数, 敗北数, 勝率(%)] を返す。
    失敗時は None を返す。"""
    if not KIF_DOWNLOAD_AVAILABLE:
        with open(log_path, "a", encoding="utf-8") as f:
            f.write("[WARN] kif_download が利用不可のため勝率集計をスキップ\n")
        return None

    try:
        # trg_YYYYMMDD_HHMM を作成し rate_trg_YYYYMMDD_HHMM にリネーム
        run_dir = kif_download.create_run_directory(ROOT, log_path)
        run_dirname = os.path.basename(run_dir)
        rate_dirname = "rate_" + run_dirname  # rate_trg_YYYYMMDD_HHMM
        rate_dir = os.path.join(os.path.dirname(run_dir), rate_dirname)
        os.rename(run_dir, rate_dir)
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"[rate] ディレクトリ作成: {rate_dir}\n")

        # 100件ダウンロード
        battles = kif_download.fetch_battles_raw(ROOT, log_path, wars_id=WARS_ID, count=100)
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"[rate] 取得した対局数: {len(battles)}\n")

        for b in battles:
            try:
                kif_text = kif_download.fetch_kif(ROOT, log_path, b["show_path"])
                kif_download.save_raw_kif(rate_dir, b["battle_id"], kif_text)
            except Exception as ex:
                with open(log_path, "a", encoding="utf-8") as f:
                    f.write(f"[WARN] fetch_kif 失敗: {b['battle_id']}: {ex}\n")

        # 1ヶ月より前のデータを削除
        cutoff = datetime.now() - timedelta(days=30)
        deleted = 0
        for battle_id in list(os.listdir(rate_dir)):
            battle_path = os.path.join(rate_dir, battle_id)
            if not os.path.isdir(battle_path):
                continue
            m = re.search(r'(\d{8})_\d{6}$', battle_id)
            if not m:
                continue
            battle_date = datetime.strptime(m.group(1), "%Y%m%d")
            if battle_date < cutoff:
                shutil.rmtree(battle_path)
                deleted += 1
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"[rate] 1ヶ月以前のデータを削除: {deleted} 件\n")

        # 5.3 追加アップデート: 以前の rate_trg ディレクトリと重複する battle_id を削除し、旧 rate_trg を削除
        temp_root = os.path.join(ROOT, "temp")
        old_rate_dirs = []
        if os.path.exists(temp_root):
            for d in os.listdir(temp_root):
                d_path = os.path.join(temp_root, d)
                if d.startswith("rate_trg_") and os.path.isdir(d_path) and d_path != rate_dir:
                    old_rate_dirs.append(d_path)

        # 既存 rate_trg の battle_id セットを構築
        existing_rate_ids = set()
        for old_dir in old_rate_dirs:
            for bid in os.listdir(old_dir):
                if os.path.isdir(os.path.join(old_dir, bid)):
                    existing_rate_ids.add(bid)

        # 重複する battle_id を削除
        dup_deleted = 0
        for battle_id in list(os.listdir(rate_dir)):
            battle_path = os.path.join(rate_dir, battle_id)
            if not os.path.isdir(battle_path):
                continue
            if battle_id in existing_rate_ids:
                shutil.rmtree(battle_path)
                dup_deleted += 1
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"[rate] 過去のrate_trgと重複するデータを削除: {dup_deleted} 件\n")
        print(f"[rate] 過去のrate_trgと重複するデータを削除: {dup_deleted} 件")

        # 以前の rate_trg ディレクトリを削除
        for old_dir in old_rate_dirs:
            shutil.rmtree(old_dir)
            with open(log_path, "a", encoding="utf-8") as f:
                f.write(f"[rate] 過去のrate_trgを削除: {os.path.basename(old_dir)}\n")
            print(f"[rate] 過去のrate_trgを削除: {os.path.basename(old_dir)}")

        # 残ったデータで勝敗を集計
        total = 0
        win_count = 0
        lose_count = 0
        for battle_id in os.listdir(rate_dir):
            battle_path = os.path.join(rate_dir, battle_id)
            if not os.path.isdir(battle_path):
                continue
            kif_path = os.path.join(battle_path, "raw.kif")
            if not os.path.exists(kif_path):
                continue
            with open(kif_path, "r", encoding="utf-8") as f:
                kif_text = f.read()
            total += 1
            if _is_user_lost(kif_text):
                lose_count += 1
            else:
                win_count += 1

        rate = round(win_count / total * 100, 1) if total > 0 else 0.0
        # 集計日付は rate_trg_YYYYMMDD_HHMM の YYYYMMDD_HHMM 部分
        timestamp = rate_dirname.replace("rate_trg_", "")
        row = [timestamp, total, win_count, lose_count, rate]

        with open(log_path, "a", encoding="utf-8") as f:
            f.write(
                f"[rate] 集計結果: {timestamp} "
                f"対局数={total}, 勝利={win_count}, 敗北={lose_count}, 勝率={rate}%\n"
            )
        return row

    except Exception:
        import traceback
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"[WARN] 勝率集計でエラー:\n{traceback.format_exc()}\n")
        return None


def run(ROOT, log_path):
    with open(log_path, "a", encoding="utf-8") as f:
        f.write("=== Step5: make_xlsx 開始 ===\n")
    print("=== Step5: make_xlsx 開始 ===")

    if not OPENPYXL_AVAILABLE:
        msg = "[SKIP] openpyxl がインストールされていないためスキップ\n"
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(msg)
        print(msg.strip())
        return

    output_path = os.path.join(ROOT, "reports", "将棋の敗因傾向.xlsx")

    try:
        # ---- データ読み込み ----
        local_summarys_path = os.path.join(ROOT, "temp", "local_summarys.json")
        local_summarys = {}
        if os.path.exists(local_summarys_path):
            with open(local_summarys_path, "r", encoding="utf-8") as f:
                local_summarys = json.load(f)
        else:
            with open(log_path, "a", encoding="utf-8") as f:
                f.write("[WARN] local_summarys.json が存在しません\n")

        sorting_path = os.path.join(ROOT, "reports", "Sorting.json")
        sorting = {}
        if os.path.exists(sorting_path):
            with open(sorting_path, "r", encoding="utf-8") as f:
                sorting = json.load(f)
        else:
            with open(log_path, "a", encoding="utf-8") as f:
                f.write("[WARN] Sorting.json が存在しません\n")

        # シート1に使う全 battle_id の raw.kif パスを収集
        all_kif_paths = _collect_battle_kif_paths(ROOT)
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"収集した battle_id 数: {len(all_kif_paths)}\n")

        # ---- 既存の勝率推移表データを読み込む（追記用） ----
        existing_rate_rows = []
        if os.path.exists(output_path):
            try:
                wb_existing = openpyxl.load_workbook(output_path)
                if "勝率推移表" in wb_existing.sheetnames:
                    ws_existing = wb_existing["勝率推移表"]
                    for row in ws_existing.iter_rows(min_row=2, values_only=True):
                        if any(cell is not None for cell in row):
                            existing_rate_rows.append(list(row))
                wb_existing.close()
                with open(log_path, "a", encoding="utf-8") as f:
                    f.write(f"既存の勝率推移データ: {len(existing_rate_rows)} 行\n")
            except Exception as ex:
                with open(log_path, "a", encoding="utf-8") as f:
                    f.write(f"[WARN] 既存 xlsx 読み込みエラー: {ex}\n")

        # ---- 今回の勝率集計（rate_trg ディレクトリ作成→DL→集計） ----
        new_rate_row = _build_rate_row(ROOT, log_path)

        # ---- Workbook 構築 ----
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)

        # ====================================================
        # シート1: 戦型別頻度表（毎回上書き）
        # ====================================================
        ws1 = wb.create_sheet("戦型別頻度表")
        style_count = defaultdict(int)

        for battle_id in local_summarys.get("各対局の振り返り", {}).keys():
            kif_path = all_kif_paths.get(battle_id)
            if kif_path and os.path.exists(kif_path):
                with open(kif_path, "r", encoding="utf-8") as f:
                    kif_text = f.read()
                style = _get_opponent_style(kif_text)
            else:
                style = "その他"
                with open(log_path, "a", encoding="utf-8") as f:
                    f.write(f"[WARN] raw.kif が見つからない: {battle_id}\n")
            style_count[style] += 1

        ws1.append(["相手の戦型", "負け回数"])
        for style, count in sorted(style_count.items(), key=lambda x: -x[1]):
            ws1.append([style, count])

        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"[シート1] 戦型別頻度表: {len(style_count)} 種類\n")
        print(f"[シート1] 戦型別頻度表: {len(style_count)} 種類")

        # ====================================================
        # シート2: 敗因別頻度表（毎回上書き）
        # ====================================================
        ws2 = wb.create_sheet("敗因別頻度表")
        themes = sorting.get("themes", {})

        ws2.append(["敗因テーマ", "負け回数"])
        for theme_name, theme_data in sorted(
            themes.items(), key=lambda x: -len(x[1].get("games", []))
        ):
            count = len(theme_data.get("games", []))
            ws2.append([theme_name, count])

        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"[シート2] 敗因別頻度表: {len(themes)} テーマ\n")
        print(f"[シート2] 敗因別頻度表: {len(themes)} テーマ")

        # ====================================================
        # シート3: 勝率推移表（追記方式: 新しい行を先頭に）
        # ====================================================
        ws3 = wb.create_sheet("勝率推移表")
        ws3.append(["集計日付", "対局数", "勝利数", "敗北数", "勝率(%)"])

        if new_rate_row:
            ws3.append(new_rate_row)
        for row in existing_rate_rows:
            ws3.append(row)

        total_rate_rows = (1 if new_rate_row else 0) + len(existing_rate_rows)
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"[シート3] 勝率推移表: 合計 {total_rate_rows} 行\n")
        print(f"[シート3] 勝率推移表: 合計 {total_rate_rows} 行")

        # ====================================================
        # シート4: 勝率推移グラフ（ws3 のデータを参照）
        # ====================================================
        ws4 = wb.create_sheet("勝率推移グラフ")

        data_rows = ws3.max_row - 1  # ヘッダー行を除いたデータ行数
        if data_rows > 0:
            chart = LineChart()
            chart.title = "勝率推移"
            chart.style = 10
            chart.y_axis.title = "勝率(%)"
            chart.x_axis.title = "集計日付"
            chart.width = 20
            chart.height = 12

            # 勝率列（E列=5列目）min_row=1 でヘッダー込みにして titles_from_data=True
            data_ref = Reference(
                ws3, min_col=5, max_col=5, min_row=1, max_row=ws3.max_row
            )
            cats_ref = Reference(
                ws3, min_col=1, max_col=1, min_row=2, max_row=ws3.max_row
            )
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            ws4.add_chart(chart, "A1")

        with open(log_path, "a", encoding="utf-8") as f:
            f.write("[シート4] 勝率推移グラフ作成完了\n")
        print("[シート4] 勝率推移グラフ作成完了")

        # ---- 保存 ----
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)

        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"[完了] 将棋の敗因傾向.xlsx 保存: {output_path}\n")
        print(f"[完了] 将棋の敗因傾向.xlsx 保存: {output_path}")

    except Exception:
        import traceback
        msg = f"[ERROR] make_xlsx.run() でエラー:\n{traceback.format_exc()}\n"
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(msg)
        print(msg)

    with open(log_path, "a", encoding="utf-8") as f:
        f.write("=== Step5: make_xlsx 完了 ===\n")
    print("=== Step5: make_xlsx 完了 ===")


if __name__ == "__main__":
    import sys
    ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    log_path = os.path.join(ROOT, "logs", f"test_make_xlsx_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
    os.makedirs(os.path.dirname(log_path), exist_ok=True)
    run(ROOT, log_path)
